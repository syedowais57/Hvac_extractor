"""
Compare Original Template vs Generated Excel
"""
import openpyxl

# Load both files
original = openpyxl.load_workbook(r'D:\SW\new project\Boeing Arlington R&D Setup.xlsx')
generated = openpyxl.load_workbook(r'D:\SW\new project\output\hvac_report.xlsx')

print('=' * 60)
print('COMPARISON: Original Template vs Generated Excel')
print('=' * 60)

print(f'\nOriginal template sheets: {len(original.sheetnames)}')
print(f'Generated Excel sheets: {len(generated.sheetnames)}')

# Compare VAV sheets
orig_vavs = [s for s in original.sheetnames if s.startswith('VAVB')]
gen_vavs = [s for s in generated.sheetnames if s.startswith('VAVB')]
print(f'\nVAV sheets - Original: {len(orig_vavs)}, Generated: {len(gen_vavs)}')

# Compare a sample VAV sheet
sample = 'VAVB5-01'
if sample in original.sheetnames and sample in generated.sheetnames:
    print(f'\n{"="*60}')
    print(f'Sample comparison: {sample}')
    print('='*60)
    
    orig_ws = original[sample]
    gen_ws = generated[sample]
    
    print('\n--- ORIGINAL TEMPLATE ---')
    for i in range(1, 25):
        row_data = []
        for j in range(1, 6):
            val = orig_ws.cell(i, j).value
            if val:
                row_data.append(str(val)[:40])
        if row_data:
            print(f'  Row {i:2d}: {" | ".join(row_data)}')
    
    print('\n--- GENERATED EXCEL ---')
    for i in range(1, 20):
        row_data = []
        for j in range(1, 5):
            val = gen_ws.cell(i, j).value
            if val:
                row_data.append(str(val)[:40])
        if row_data:
            print(f'  Row {i:2d}: {" | ".join(row_data)}')

# Check data values
print('\n' + '='*60)
print('DATA VALUES COMPARISON')
print('='*60)

# Get CFM from original (look in column N which is column 14)
orig_cfm = None
for i in range(1, 30):
    val = orig_ws.cell(i, 14).value  # Column N
    if val and isinstance(val, (int, float)) and val > 100:
        orig_cfm = val
        print(f'Original CFM (row {i}, col N): {val}')
        break

# Get CFM from generated
for i in range(1, 20):
    for j in range(1, 5):
        cell_val = gen_ws.cell(i, j).value
        label = gen_ws.cell(i, 1).value
        if label and 'Maximum CFM' in str(label):
            gen_cfm = gen_ws.cell(i, 2).value
            print(f'Generated CFM (row {i}): {gen_cfm}')
            break
