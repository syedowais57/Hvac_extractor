"""
Detailed comparison between original and generated Excel files.
Creates an Excel report with missing sheets and values.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load workbooks
original = openpyxl.load_workbook(r'D:\SW\new project\Boeing Arlington R&D Setup.xlsx')
generated = openpyxl.load_workbook(r'D:\SW\new project\output\hvac_report.xlsx')

# Create report workbook
report = openpyxl.Workbook()
report.remove(report.active)

# Styles
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
missing_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
present_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# SHEET 1: Summary
# ============================================================
summary_ws = report.create_sheet("Summary")

summary_ws['A1'] = "EXCEL COMPARISON REPORT"
summary_ws['A1'].font = Font(bold=True, size=16)
summary_ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

summary_ws['A4'] = "Original File:"
summary_ws['B4'] = "Boeing Arlington R&D Setup.xlsx"
summary_ws['A5'] = "Generated File:"
summary_ws['B5'] = "hvac_report.xlsx"

summary_ws['A7'] = "Metric"
summary_ws['B7'] = "Original"
summary_ws['C7'] = "Generated"
summary_ws['D7'] = "Difference"

for col in ['A', 'B', 'C', 'D']:
    summary_ws[f'{col}7'].font = header_font
    summary_ws[f'{col}7'].fill = header_fill

orig_sheets = set(original.sheetnames)
gen_sheets = set(generated.sheetnames)
missing_sheets = orig_sheets - gen_sheets
extra_sheets = gen_sheets - orig_sheets

summary_ws['A8'] = "Total Sheets"
summary_ws['B8'] = len(orig_sheets)
summary_ws['C8'] = len(gen_sheets)
summary_ws['D8'] = len(gen_sheets) - len(orig_sheets)

summary_ws['A9'] = "Missing Sheets"
summary_ws['B9'] = "-"
summary_ws['C9'] = len(missing_sheets)
summary_ws['D9'] = f"{len(missing_sheets)} not generated"

summary_ws['A10'] = "Extra Sheets"
summary_ws['B10'] = "-"
summary_ws['C10'] = len(extra_sheets)
summary_ws['D10'] = f"{len(extra_sheets)} new"

# ============================================================
# SHEET 2: Missing Sheets
# ============================================================
missing_ws = report.create_sheet("Missing Sheets")

missing_ws['A1'] = "MISSING SHEETS (in original but not in generated)"
missing_ws['A1'].font = Font(bold=True, size=14)

# Categorize missing sheets
categories = {
    'VAV': [],
    'EF (Exhaust Fans)': [],
    'Electric Duct Heaters': [],
    'Flow Meters': [],
    'Other': []
}

for s in sorted(missing_sheets):
    if s.startswith('VAVB'):
        categories['VAV'].append(s)
    elif s.startswith('EF'):
        categories['EF (Exhaust Fans)'].append(s)
    elif 'Heater' in s or 'Electric' in s:
        categories['Electric Duct Heaters'].append(s)
    elif 'Flow' in s:
        categories['Flow Meters'].append(s)
    else:
        categories['Other'].append(s)

row = 3
for cat, sheets in categories.items():
    if sheets:
        missing_ws[f'A{row}'] = f"{cat} ({len(sheets)} sheets)"
        missing_ws[f'A{row}'].font = Font(bold=True)
        missing_ws[f'A{row}'].fill = header_fill
        missing_ws[f'A{row}'].font = header_font
        row += 1
        for s in sheets:
            missing_ws[f'A{row}'] = s
            missing_ws[f'A{row}'].fill = missing_fill
            row += 1
        row += 1

# ============================================================
# SHEET 3: Extra Sheets (Generated but not in Original)
# ============================================================
extra_ws = report.create_sheet("Extra Sheets (Generated)")

extra_ws['A1'] = "EXTRA SHEETS (in generated but NOT in original)"
extra_ws['A1'].font = Font(bold=True, size=14)

extra_ws['A3'] = "Sheet Name"
extra_ws['B3'] = "Notes"
extra_ws['A3'].font = header_font
extra_ws['A3'].fill = header_fill
extra_ws['B3'].font = header_font
extra_ws['B3'].fill = header_fill

row = 4
if extra_sheets:
    for s in sorted(extra_sheets):
        extra_ws[f'A{row}'] = s
        extra_ws[f'A{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        extra_ws[f'B{row}'] = "Generated but not in original template"
        row += 1
else:
    extra_ws[f'A{row}'] = "(No extra sheets)"
    
# ============================================================
# SHEET 4: Sheet by Sheet Comparison
# ============================================================
compare_ws = report.create_sheet("Sheet Comparison")

compare_ws['A1'] = "SHEET BY SHEET STATUS"
compare_ws['A1'].font = Font(bold=True, size=14)

compare_ws['A3'] = "Sheet Name"
compare_ws['B3'] = "In Original"
compare_ws['C3'] = "In Generated"
compare_ws['D3'] = "Status"

for col in ['A', 'B', 'C', 'D']:
    compare_ws[f'{col}3'].font = header_font
    compare_ws[f'{col}3'].fill = header_fill

row = 4
all_sheets = sorted(orig_sheets | gen_sheets)
for sheet in all_sheets:
    in_orig = sheet in orig_sheets
    in_gen = sheet in gen_sheets
    
    compare_ws[f'A{row}'] = sheet
    compare_ws[f'B{row}'] = "Yes" if in_orig else "No"
    compare_ws[f'C{row}'] = "Yes" if in_gen else "No"
    
    if in_orig and in_gen:
        compare_ws[f'D{row}'] = "✓ Matched"
        compare_ws[f'D{row}'].fill = present_fill
    elif in_orig and not in_gen:
        compare_ws[f'D{row}'] = "✗ Missing"
        compare_ws[f'D{row}'].fill = missing_fill
    else:
        compare_ws[f'D{row}'] = "+ New"
        compare_ws[f'D{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    row += 1

# ============================================================
# SHEET 4: VAV Value Comparison (for common sheets)
# ============================================================
vav_ws = report.create_sheet("VAV Value Comparison")

vav_ws['A1'] = "VAV SHEET VALUE COMPARISON"
vav_ws['A1'].font = Font(bold=True, size=14)

# Find common VAV sheets
common_vav = [s for s in orig_sheets if s.startswith('VAVB') and s in gen_sheets]

vav_ws['A3'] = "Field"
vav_ws['B3'] = "Original Value"
vav_ws['C3'] = "Generated Value"
vav_ws['D3'] = "Match"

for col in ['A', 'B', 'C', 'D']:
    vav_ws[f'{col}3'].font = header_font
    vav_ws[f'{col}3'].fill = header_fill

row = 4

# Key fields to compare (based on typical VAV sheet structure)
key_fields = [
    ("Tag", "VAVB ID"),
    ("Area Served", "Area"),
    ("Room", "Location"),
    ("Max CFM", "Max CFM"),
    ("Min CFM", "Min CFM"),
    ("Manufacturer", "Manufacturer"),
    ("Model Number", "Model"),
]

# Compare first common VAV sheet as example
if common_vav:
    sample_sheet = common_vav[0]
    vav_ws[f'A{row}'] = f"Sample Sheet: {sample_sheet}"
    vav_ws[f'A{row}'].font = Font(bold=True, italic=True)
    row += 2
    
    orig_ws = original[sample_sheet]
    gen_ws = generated[sample_sheet]
    
    # Extract all values from both sheets
    def get_sheet_values(ws, max_row=40, max_col=6):
        values = {}
        for r in range(1, max_row):
            for c in range(1, max_col):
                val = ws.cell(r, c).value
                if val:
                    values[(r, c)] = val
        return values
    
    orig_values = get_sheet_values(orig_ws)
    gen_values = get_sheet_values(gen_ws)
    
    # Find label:value pairs in original
    vav_ws[f'A{row}'] = "Original Sheet Structure:"
    vav_ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    for (r, c), val in sorted(orig_values.items()):
        vav_ws[f'A{row}'] = f"Row {r}, Col {c}"
        vav_ws[f'B{row}'] = str(val)[:50] if val else ""
        row += 1
        if row > 50:
            break
    
    row += 2
    vav_ws[f'A{row}'] = "Generated Sheet Structure:"
    vav_ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    for (r, c), val in sorted(gen_values.items()):
        vav_ws[f'A{row}'] = f"Row {r}, Col {c}"
        vav_ws[f'B{row}'] = str(val)[:50] if val else ""
        row += 1
        if row > 100:
            break

# ============================================================
# SHEET 5: Missing Values Detail
# ============================================================
missing_val_ws = report.create_sheet("Missing Values")

missing_val_ws['A1'] = "MISSING VALUES IN GENERATED SHEETS"
missing_val_ws['A1'].font = Font(bold=True, size=14)

row = 3
for sheet_name in common_vav[:5]:  # Check first 5 common VAV sheets
    orig_ws = original[sheet_name]
    gen_ws = generated[sheet_name]
    
    missing_val_ws[f'A{row}'] = f"Sheet: {sheet_name}"
    missing_val_ws[f'A{row}'].font = Font(bold=True)
    missing_val_ws[f'A{row}'].fill = header_fill
    missing_val_ws[f'A{row}'].font = header_font
    row += 1
    
    # Get all non-empty cells from original
    orig_data = {}
    for r in range(1, 40):
        for c in range(1, 7):
            val = orig_ws.cell(r, c).value
            if val:
                orig_data[(r, c)] = val
    
    # Get all non-empty cells from generated
    gen_data = {}
    for r in range(1, 30):
        for c in range(1, 5):
            val = gen_ws.cell(r, c).value
            if val:
                gen_data[(r, c)] = val
    
    # Find values in original that don't appear anywhere in generated
    orig_text_values = set(str(v).strip().lower() for v in orig_data.values() if v)
    gen_text_values = set(str(v).strip().lower() for v in gen_data.values() if v)
    
    missing_values = orig_text_values - gen_text_values
    
    if missing_values:
        for val in sorted(missing_values):
            if len(val) > 2:  # Skip very short values
                missing_val_ws[f'A{row}'] = val
                missing_val_ws[f'A{row}'].fill = missing_fill
                row += 1
    else:
        missing_val_ws[f'A{row}'] = "(No missing values)"
        row += 1
    
    row += 1

# ============================================================
# Adjust column widths
# ============================================================
for ws in report.worksheets:
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 25

# Save report
report_path = r'D:\SW\new project\output\comparison_report.xlsx'
report.save(report_path)
print(f"✓ Comparison report saved to: {report_path}")

# Print summary to console
print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"Original sheets: {len(orig_sheets)}")
print(f"Generated sheets: {len(gen_sheets)}")
print(f"Missing sheets: {len(missing_sheets)}")
print(f"Extra sheets (generated but not in original): {len(extra_sheets)}")
print(f"\nCommon VAV sheets found: {len(common_vav)}")
print("\nMissing sheet categories:")
for cat, sheets in categories.items():
    if sheets:
        print(f"  {cat}: {len(sheets)}")

if extra_sheets:
    print("\nExtra sheets (in generated but NOT in original):")
    for s in sorted(extra_sheets):
        print(f"  + {s}")
