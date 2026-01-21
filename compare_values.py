"""
Compare extracted values with original Excel sheet values
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

# Load workbooks
original = openpyxl.load_workbook(r'D:\SW\new project\Boeing Arlington R&D Setup.xlsx')
generated = openpyxl.load_workbook(r'D:\SW\new project\output\hvac_report_detailed.xlsx')

# Find common VAV sheets
orig_vavs = [s for s in original.sheetnames if s.startswith('VAVB')]
gen_vavs = [s for s in generated.sheetnames if s.startswith('VAVB')]
common_vavs = sorted(set(orig_vavs) & set(gen_vavs))

print("=" * 70)
print("VALUE COMPARISON: Original vs Generated")
print("=" * 70)
print(f"Common VAV sheets to compare: {len(common_vavs)}")
print()

# Fields to compare (label in original -> approximate row)
# We'll search for these labels and compare values

def get_cell_value_by_label(ws, label, value_col=None, search_rows=40, search_cols=12):
    """Find a label and return the value in the specified column"""
    for r in range(1, search_rows):
        label_cell = ws.cell(r, 1).value  # Labels are typically in column A (1)
        if label_cell and str(label_cell).strip().lower() == label.lower():
            # If value_col provided, use it, otherwise use next column
            col = value_col if value_col else 2
            return ws.cell(r, col).value
    return None

# Compare VAV sheets in detail
fields_to_compare = [
    "Unit Number",
    "Location", 
    "Area Served",
    "Manufacturer",
    "Model Number",
    "Primary Air Inlet Size",
    "Total Fan CFM",
    "Minimum CFM",
    "Maximum CFM",
    "Motor HP",
    "Motor Voltage",
    "Reheat KW",
]

results = []

print(f"\n{'='*90}")
print(f"{'VAV Tag':<10} {'Field':<25} {'Original':<20} {'Generated':<20} {'Match'}")
print(f"{'='*90}")

for vav_tag in common_vavs[:15]:  # Compare first 15
    orig_ws = original[vav_tag]
    gen_ws = generated[vav_tag]
    
    for field in fields_to_compare:
        # Original: labels in A, values in K(11)
        orig_val = get_cell_value_by_label(orig_ws, field, value_col=11)
        # Generated: labels in A, values in B(2)
        gen_val = get_cell_value_by_label(gen_ws, field, value_col=2)
        
        if orig_val is None and gen_val is None:
            continue
            
        # Normalize for comparison
        orig_str = str(orig_val).strip().replace('"', '').replace('None', '') if orig_val is not None else ""
        gen_str = str(gen_val).strip().replace('"', '').replace('None', '') if gen_val is not None else ""
        
        # Check if match
        try:
            # Try numeric comparison
            o_num = float(orig_str.split()[0]) if orig_str else 0
            g_num = float(gen_str.split()[0]) if gen_str else 0
            match = o_num == g_num
        except:
            match = orig_str.lower() == gen_str.lower()
        
        status = "✓" if match else "✗"
        
        # Determine mismatch type for reporting
        if not match:
            print(f"{vav_tag:<10} {field:<25} {orig_str[:18]:<20} {gen_str[:18]:<20} {status}")
            
    results.append(vav_tag)

print("\n" + "=" * 90)
print("NOTE: Only showing MISMATCHES above for clarity.")
print("=" * 90)
print(f"Total VAVs compared: {len(results)}")

