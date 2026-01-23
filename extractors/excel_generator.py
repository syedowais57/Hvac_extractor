"""
Excel Generator for HVAC Data
Creates a NEW Excel workbook from scratch with extracted HVAC data
Similar structure to Boeing template but generated programmatically
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from pathlib import Path
from datetime import datetime
import json


class HVACExcelGenerator:
    """
    Generates a new Excel workbook from scratch with HVAC data
    Creates professional-looking sheets similar to industry templates
    """
    
    # Styles
    HEADER_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, size=11)
    LABEL_FONT = Font(bold=True)
    DATA_FONT = Font()
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    def __init__(self, job_number: str = "1168", project_name: str = "HVAC Project"):
        self.wb = openpyxl.Workbook()
        self.job_number = job_number
        self.project_name = project_name
        self.date = datetime.now().strftime("%Y-%m-%d")
        
    def _set_column_widths(self, ws, widths: Dict[str, int]):
        """Set column widths"""
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    def _add_header(self, ws, title: str, row: int = 1):
        """Add sheet header"""
        ws.cell(row=row, column=1, value=title).font = self.HEADER_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
    def _add_job_info(self, ws, system_name: str, start_row: int = 3):
        """Add job information section"""
        ws.cell(row=start_row, column=1, value="Job Number:").font = self.LABEL_FONT
        ws.cell(row=start_row, column=2, value=self.job_number)
        ws.cell(row=start_row, column=3, value="Date:").font = self.LABEL_FONT
        ws.cell(row=start_row, column=4, value=self.date)
        
        ws.cell(row=start_row + 1, column=1, value="System:").font = self.LABEL_FONT
        ws.cell(row=start_row + 1, column=2, value=system_name)
        ws.cell(row=start_row + 1, column=3, value="Project:").font = self.LABEL_FONT
        ws.cell(row=start_row + 1, column=4, value=self.project_name)
        
    def _add_section_header(self, ws, title: str, row: int):
        """Add a section header with background"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.SECTION_FONT
        cell.fill = self.SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
    def create_vav_sheet(self, vav_data: Dict) -> str:
        """Create a VAV box test data sheet with all extracted details"""
        tag = vav_data.get("tag", "VAV-XX")
        ws = self.wb.create_sheet(title=tag)
        
        # Column widths - more columns for complete data
        self._set_column_widths(ws, {"A": 25, "B": 20, "C": 15, "D": 15, "E": 15})
        
        # Header
        self._add_header(ws, "Fan Powered VAV Box Test Data")
        self._add_job_info(ws, tag, start_row=3)
        
        # Unit Information Section
        self._add_section_header(ws, "UNIT INFORMATION", 6)
        
        unit_info = [
            ("Unit Number", tag),
            ("Location", vav_data.get("location") or ""),
            ("Area Served", vav_data.get("area_served") or ""),
            ("Manufacturer", vav_data.get("manufacturer") or ""),
            ("Model Number", vav_data.get("model") or ""),
            ("Primary Air Inlet Size", f'{vav_data.get("inlet_size", "")}\"' if vav_data.get("inlet_size") else ""),
        ]
        
        row = 7
        for label, value in unit_info:
            ws.cell(row=row, column=1, value=label).font = self.LABEL_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Air Measurements Section
        self._add_section_header(ws, "AIR MEASUREMENTS", row + 1)
        row += 2
        
        # Headers
        ws.cell(row=row, column=1, value="Parameter").font = self.LABEL_FONT
        ws.cell(row=row, column=2, value="DESIGN").font = self.LABEL_FONT
        ws.cell(row=row, column=3, value="ACTUAL").font = self.LABEL_FONT
        row += 1
        
        air_data = [
            ("Total Fan CFM", vav_data.get("total_cfm") or vav_data.get("cfm_max", "")),
            ("Minimum CFM", vav_data.get("cfm_min", "")),
            ("Maximum CFM", vav_data.get("cfm_max", "")),
            ("Fan Speed Setting", ""),
            ("DDC Calibration Factor", ""),
            ("DDC Address", ""),
        ]
        
        for label, value in air_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value="")  # Actual - to be filled in field
            row += 1
        
        # Motor Measurements Section
        self._add_section_header(ws, "MOTOR MEASUREMENTS", row + 1)
        row += 2
        
        ws.cell(row=row, column=1, value="Parameter").font = self.LABEL_FONT
        ws.cell(row=row, column=2, value="DESIGN").font = self.LABEL_FONT
        ws.cell(row=row, column=3, value="ACTUAL").font = self.LABEL_FONT
        row += 1
        
        motor_data = [
            ("Motor HP", vav_data.get("motor_hp", "")),
            ("Motor Voltage", vav_data.get("motor_voltage", "")),
            ("Motor Phase", vav_data.get("motor_phase", "")),
            ("Motor Amperage", vav_data.get("motor_amperage", "")),
            ("CFLA", ""),
        ]
        
        for label, value in motor_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value="")
            row += 1
        
        # Reheat Section (if applicable)
        if vav_data.get("has_reheat") or vav_data.get("reheat_kw"):
            self._add_section_header(ws, "ELECTRIC REHEAT", row + 1)
            row += 2
            
            ws.cell(row=row, column=1, value="Reheat KW").font = self.LABEL_FONT
            ws.cell(row=row, column=2, value=vav_data.get("reheat_kw", ""))
            ws.cell(row=row, column=3, value="")
            
        return tag
    
    def create_fan_sheet(self, fan_data: Dict) -> str:
        """Create an Exhaust Fan test data sheet"""
        tag = fan_data.get("tag", "EF-X")
        ws = self.wb.create_sheet(title=tag)
        
        self._set_column_widths(ws, {"A": 30, "B": 15, "C": 15, "D": 15})
        
        # Header
        self._add_header(ws, "Direct Drive Fan Test Data")
        self._add_job_info(ws, tag, start_row=3)
        
        # Unit Information
        self._add_section_header(ws, "UNIT INFORMATION", 6)
        
        unit_info = [
            ("Unit Number", tag),
            ("Location", fan_data.get("location") or ""),
            ("Type", fan_data.get("fan_type") or ""),
            ("Drive", fan_data.get("drive") or ""),
        ]
        
        row = 7
        for label, value in unit_info:
            ws.cell(row=row, column=1, value=label).font = self.LABEL_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Air Measurements
        self._add_section_header(ws, "AIR MEASUREMENTS", row + 1)
        row += 2
        
        ws.cell(row=row, column=1, value="Parameter").font = self.LABEL_FONT
        ws.cell(row=row, column=2, value="Design").font = self.LABEL_FONT
        ws.cell(row=row, column=3, value="Actual").font = self.LABEL_FONT
        row += 1
        
        air_data = [
            ("Total Fan CFM", fan_data.get("cfm", "")),
            ("External Static Pressure (in WG)", fan_data.get("esp", "")),
            ("Fan RPM", fan_data.get("rpm", "")),
        ]
        
        for label, value in air_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Motor Measurements
        self._add_section_header(ws, "MOTOR MEASUREMENTS", row + 1)
        row += 2
        
        motor_data = [
            ("Motor Power", fan_data.get("motor_power") or ""),
            ("Voltage", fan_data.get("voltage") or ""),
        ]
        
        for label, value in motor_data:
            ws.cell(row=row, column=1, value=label).font = self.LABEL_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        return tag
    
    def create_crac_sheet(self, crac_data: Dict) -> str:
        """Create a CRAC unit test data sheet"""
        tag = crac_data.get("tag", "CRAC-X")
        
        # Check if sheet already exists
        if tag in self.wb.sheetnames:
            return tag
            
        ws = self.wb.create_sheet(title=tag)
        
        self._set_column_widths(ws, {"A": 30, "B": 15, "C": 15, "D": 15})
        
        self._add_header(ws, "Computer Room AC Unit Test Data")
        self._add_job_info(ws, tag, start_row=3)
        
        # Unit Information
        self._add_section_header(ws, "UNIT INFORMATION", 6)
        
        row = 7
        unit_info = [
            ("Unit Number", tag),
            ("Location", crac_data.get("location") or ""),
        ]
        
        for label, value in unit_info:
            ws.cell(row=row, column=1, value=label).font = self.LABEL_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Performance
        self._add_section_header(ws, "PERFORMANCE", row + 1)
        row += 2
        
        ws.cell(row=row, column=1, value="Parameter").font = self.LABEL_FONT
        ws.cell(row=row, column=2, value="Design").font = self.LABEL_FONT
        ws.cell(row=row, column=3, value="Actual").font = self.LABEL_FONT
        row += 1
        
        perf_data = [
            ("CFM", crac_data.get("cfm", "")),
            ("Cooling Capacity", crac_data.get("cooling_capacity") or ""),
        ]
        
        for label, value in perf_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        return tag
    
    def create_heater_sheet(self, heater_data: Dict) -> str:
        """Create an Electric Duct Heater test data sheet"""
        tag = heater_data.get("tag", "EDH-X")
        
        # Check if sheet exists
        if tag in self.wb.sheetnames:
            return tag
            
        ws = self.wb.create_sheet(title=tag)
        
        self._set_column_widths(ws, {"A": 30, "B": 15, "C": 15, "D": 15})
        
        self._add_header(ws, "Electric Duct Heater Test Data")
        self._add_job_info(ws, tag, start_row=3)
        
        # Heater Information
        self._add_section_header(ws, "HEATER INFORMATION", 6)
        
        row = 7
        unit_info = [
            ("Unit Number", tag),
            ("Location", heater_data.get("location") or ""),
            ("Associated VAV", heater_data.get("associated_vav") or ""),
        ]
        
        for label, value in unit_info:
            ws.cell(row=row, column=1, value=label).font = self.LABEL_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Performance
        self._add_section_header(ws, "HEATER PERFORMANCE", row + 1)
        row += 2
        
        ws.cell(row=row, column=1, value="Parameter").font = self.LABEL_FONT
        ws.cell(row=row, column=2, value="Design").font = self.LABEL_FONT
        ws.cell(row=row, column=3, value="Actual").font = self.LABEL_FONT
        row += 1
        
        perf_data = [
            ("CFM", heater_data.get("cfm", "")),
            ("Voltage", heater_data.get("voltage", "")),
            ("KW", heater_data.get("kw", "")),
        ]
        
        for label, value in perf_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        return tag
    
    def create_summary_sheet(self, data: Dict[str, List]):
        """Create a summary sheet with all equipment"""
        # Use the default first sheet as summary
        ws = self.wb.active
        ws.title = "Summary"
        
        self._set_column_widths(ws, {"A": 15, "B": 15, "C": 12, "D": 12, "E": 12, "F": 12, "G": 15})
        
        self._add_header(ws, f"HVAC Equipment Summary - {self.project_name}")
        
        row = 3
        ws.cell(row=row, column=1, value=f"Job Number: {self.job_number}")
        ws.cell(row=row, column=3, value=f"Date: {self.date}")
        
        # VAV Summary
        row = 5
        ws.cell(row=row, column=1, value="VAV UNITS").font = self.SECTION_FONT
        row += 1
        
        headers = ["Tag", "Location", "Max CFM", "Min CFM", "Inlet Size", "Reheat", "Reheat KW"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.LABEL_FONT
            cell.fill = self.SECTION_FILL
        row += 1
        
        for vav in data.get("vavs", []):
            ws.cell(row=row, column=1, value=vav.get("tag", ""))
            ws.cell(row=row, column=2, value=vav.get("location") or "")
            ws.cell(row=row, column=3, value=vav.get("cfm_max", ""))
            ws.cell(row=row, column=4, value=vav.get("cfm_min", ""))
            ws.cell(row=row, column=5, value=vav.get("inlet_size", ""))
            ws.cell(row=row, column=6, value="Yes" if vav.get("has_reheat") else "No")
            ws.cell(row=row, column=7, value=vav.get("reheat_kw", ""))
            row += 1
        
        # Fans Summary
        row += 2
        ws.cell(row=row, column=1, value="EXHAUST FANS").font = self.SECTION_FONT
        row += 1
        
        headers = ["Tag", "Location", "Type", "CFM", "ESP", "RPM", "Voltage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.LABEL_FONT
            cell.fill = self.SECTION_FILL
        row += 1
        
        for fan in data.get("fans", []):
            ws.cell(row=row, column=1, value=fan.get("tag", ""))
            ws.cell(row=row, column=2, value=fan.get("location") or "")
            ws.cell(row=row, column=3, value=fan.get("fan_type") or "")
            ws.cell(row=row, column=4, value=fan.get("cfm", ""))
            ws.cell(row=row, column=5, value=fan.get("esp", ""))
            ws.cell(row=row, column=6, value=fan.get("rpm", ""))
            ws.cell(row=row, column=7, value=fan.get("voltage") or "")
            row += 1
        
        # CRAC Summary
        row += 2
        ws.cell(row=row, column=1, value="CRAC UNITS").font = self.SECTION_FONT
        row += 1
        
        headers = ["Tag", "Location", "CFM", "Cooling Capacity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.LABEL_FONT
            cell.fill = self.SECTION_FILL
        row += 1
        
        for crac in data.get("cracs", []):
            ws.cell(row=row, column=1, value=crac.get("tag", ""))
            ws.cell(row=row, column=2, value=crac.get("location") or "")
            ws.cell(row=row, column=3, value=crac.get("cfm", ""))
            ws.cell(row=row, column=4, value=crac.get("cooling_capacity") or "")
            row += 1
    
    def generate_from_data(self, data: Dict[str, List]) -> Dict[str, int]:
        """
        Generate complete workbook from extracted data
        
        Args:
            data: Dict with keys: vavs, fans, cracs, heaters, air_devices
            
        Returns:
            Dict with count of sheets created per type
        """
        stats = {"vavs": 0, "fans": 0, "cracs": 0, "heaters": 0}
        
        # Create summary first
        self.create_summary_sheet(data)
        
        # Create individual VAV sheets
        seen_vavs = set()
        for vav in data.get("vavs", []):
            tag = vav.get("tag", "")
            if tag and tag not in seen_vavs:
                self.create_vav_sheet(vav)
                seen_vavs.add(tag)
                stats["vavs"] += 1
        
        # Create fan sheets
        seen_fans = set()
        for fan in data.get("fans", []):
            tag = fan.get("tag", "")
            if tag and tag not in seen_fans:
                self.create_fan_sheet(fan)
                seen_fans.add(tag)
                stats["fans"] += 1
        
        # Create CRAC sheets
        seen_cracs = set()
        for crac in data.get("cracs", []):
            tag = crac.get("tag", "")
            if tag and tag not in seen_cracs:
                self.create_crac_sheet(crac)
                seen_cracs.add(tag)
                stats["cracs"] += 1
        
        # Create heater sheets from explicit heater data
        seen_heaters = set()
        for heater in data.get("heaters", []):
            tag = heater.get("tag", "")
            if tag and tag not in seen_heaters:
                self.create_heater_sheet(heater)
                seen_heaters.add(tag)
                stats["heaters"] += 1
        
        # Also generate heater sheets from VAVs with reheat_kw
        for vav in data.get("vavs", []):
            vav_tag = vav.get("tag", "")
            reheat_kw = vav.get("reheat_kw")
            
            # Check if VAV has reheat (either explicit flag or non-zero kW)
            has_reheat = vav.get("has_reheat") or (reheat_kw and str(reheat_kw) not in ["0", "0.0", "", "null", "None"])
            
            if vav_tag and has_reheat:
                heater_tag = f"{vav_tag}-H"
                if heater_tag not in seen_heaters:
                    heater_data = {
                        "tag": heater_tag,
                        "location": vav.get("location", ""),
                        "cfm": vav.get("cfm_max", 0),
                        "voltage": 277,
                        "kw": reheat_kw,
                        "associated_vav": vav_tag
                    }
                    self.create_heater_sheet(heater_data)
                    seen_heaters.add(heater_tag)
                    stats["heaters"] += 1
        
        # Fallback: Generate blank Electric Duct Heater template sheets if none were created
        # This matches the original template which has blank heater sheets
        if stats["heaters"] == 0:
            num_heater_templates = 10  # Match original template
            for i in range(num_heater_templates):
                suffix = "" if i == 0 else f" ({i})"
                heater_tag = f"Electric Duct Heater{suffix}"
                heater_data = {
                    "tag": heater_tag,
                    "location": "",
                    "cfm": "",
                    "voltage": "",
                    "kw": "",
                    "associated_vav": ""
                }
                self.create_heater_sheet(heater_data)
                stats["heaters"] += 1
        
        return stats
    
    def save(self, output_path: str):
        """Save the workbook"""
        self.wb.save(output_path)
        print(f"[OK] Saved to: {output_path}")
        return output_path


def generate_excel_from_json(json_path: str, output_path: str, 
                              job_number: str = "1168", 
                              project_name: str = "HVAC Project") -> str:
    """
    Generate Excel from extracted JSON data
    
    Args:
        json_path: Path to extracted JSON
        output_path: Output Excel path
        job_number: Job number for the report
        project_name: Project name
        
    Returns:
        Path to generated Excel
    """
    with open(json_path) as f:
        data = json.load(f)
    
    generator = HVACExcelGenerator(job_number=job_number, project_name=project_name)
    stats = generator.generate_from_data(data)
    
    print(f"\nGeneration Summary:")
    print(f"  VAV sheets: {stats['vavs']}")
    print(f"  Fan sheets: {stats['fans']}")
    print(f"  CRAC sheets: {stats['cracs']}")
    print(f"  Heater sheets: {stats['heaters']}")
    print(f"  Total sheets: {sum(stats.values()) + 1}")  # +1 for summary
    
    return generator.save(output_path)


if __name__ == "__main__":
    JSON_DATA = r"D:\SW\new project\Boeing R&D Drawings_extracted.json"
    OUTPUT = r"D:\SW\new project\output\hvac_generated.xlsx"
    
    if Path(JSON_DATA).exists():
        Path(OUTPUT).parent.mkdir(parents=True, exist_ok=True)
        output = generate_excel_from_json(
            JSON_DATA, 
            OUTPUT,
            job_number="1168",
            project_name="Boeing Arlington R&D"
        )
        print(f"\nOutput: {output}")
    else:
        print(f"No extracted data found at {JSON_DATA}")
