"""
Excel Populator for HVAC Data
Populates the Boeing template Excel with extracted HVAC data
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from typing import Dict, List, Any
from pathlib import Path
import json


class HVACExcelPopulator:
    """
    Populates Boeing Arlington R&D Setup Excel template with extracted HVAC data
    """
    
    # Sheet structure mappings based on Excel analysis
    VAV_MAPPING = {
        # Row: (label, data_column)
        8: ("Unit Number", "K"),
        9: ("Location", "K"),
        10: ("Area Served", "K"),
        11: ("Manufacturer", "K"),
        12: ("Model Number", "K"),
        13: ("Inlet Size", "K"),
        16: ("Total CFM", "K"),
        17: ("Minimum CFM", "K"),
        18: ("Maximum CFM", "K"),
        24: ("Motor HP", "K"),
        25: ("Motor Voltage", "K"),
        26: ("Motor Phase", "K"),
        27: ("Motor Amperage", "K"),
    }
    
    EF_MAPPING = {
        8: ("Unit Number", "N"),
        9: ("Location", "N"),
        10: ("Area Served", "N"),
        16: ("Total Fan CFM", "N"),
        18: ("External Fan Static Pressure", "N"),
        21: ("Fan RPM", "N"),
        26: ("Motor Voltage", "N"),
        27: ("Motor Phase", "N"),
    }
    
    HEATER_MAPPING = {
        # First heater block (rows 8-20)
        "block1": {
            8: ("Unit Number", "K"),
            9: ("Location", "K"),
            14: ("CFM", "K"),
            17: ("Voltage", "K"),
            19: ("KW", "K"),
        },
        # Second heater block (rows 23-36)
        "block2": {
            24: ("Unit Number", "K"),
            25: ("Location", "K"),
            30: ("CFM", "K"),
            33: ("Voltage", "K"),
            35: ("KW", "K"),
        }
    }
    
    def __init__(self, template_path: str):
        """Load the Excel template"""
        self.template_path = template_path
        self.wb = openpyxl.load_workbook(template_path)
        self.sheet_names = self.wb.sheetnames
    
    def _safe_set_cell(self, ws, row: int, column: int, value):
        """Safely set cell value, handling merged cells"""
        try:
            cell = ws.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                # Find the top-left cell of the merged range
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                        top_left.value = value
                        return True
            else:
                cell.value = value
                return True
        except Exception as e:
            print(f"    Warning: Could not set cell ({row}, {column}): {e}")
            return False
        return False
        
    def _find_sheet_for_tag(self, tag: str, prefix: str = None) -> str:
        """Find the sheet name that matches an equipment tag"""
        # Direct match
        if tag in self.sheet_names:
            return tag
        
        # Try with prefix variations
        for sheet in self.sheet_names:
            if sheet.upper() == tag.upper():
                return sheet
            if prefix and sheet.startswith(prefix):
                if tag in sheet:
                    return sheet
        
        return None
    
    def populate_vav(self, vav_data: Dict) -> bool:
        """Populate a VAV sheet with extracted data"""
        tag = vav_data.get("tag", "")
        sheet_name = self._find_sheet_for_tag(tag, "VAV")
        
        if not sheet_name:
            print(f"  No sheet found for VAV: {tag}")
            return False
        
        ws = self.wb[sheet_name]
        
        # Map data to cells
        data_map = {
            8: tag,  # Unit Number
            9: vav_data.get("location", ""),
            10: vav_data.get("area_served") or vav_data.get("location", ""),
            11: vav_data.get("manufacturer", ""),
            12: vav_data.get("model", ""),
            13: vav_data.get("inlet_size", ""),
            16: vav_data.get("total_cfm") or vav_data.get("cfm_max", ""),
            17: vav_data.get("cfm_min", ""),
            18: vav_data.get("cfm_max", ""),
            24: vav_data.get("motor_hp", ""),
            25: vav_data.get("motor_voltage", ""),
            26: vav_data.get("motor_phase", ""),
            27: vav_data.get("motor_amperage", ""),
        }
        
        for row, value in data_map.items():
            if value is not None and value != "":
                self._safe_set_cell(ws, row, 11, value)  # Column K
        
        # Check for reheat
        if vav_data.get("has_reheat"):
            self._safe_set_cell(ws, 20, 11, vav_data.get("reheat_kw", "")) # Guessing Row 20 in Col K
        
        print(f"  [OK] Populated VAV sheet: {sheet_name}")
        return True
    
    def populate_ef(self, ef_data: Dict) -> bool:
        """Populate an Exhaust Fan sheet"""
        tag = ef_data.get("tag", "")
        sheet_name = self._find_sheet_for_tag(tag, "EF")
        
        if not sheet_name:
            print(f"  No sheet found for EF: {tag}")
            return False
        
        ws = self.wb[sheet_name]
        
        # Handle voltage safely - could be None or have "/" separator
        voltage_raw = ef_data.get("voltage")
        voltage = ""
        if voltage_raw:
            voltage = str(voltage_raw).split("/")[0] if "/" in str(voltage_raw) else str(voltage_raw)
        
        data_map = {
            8: tag,
            9: ef_data.get("location") or "",
            10: ef_data.get("location") or "",
            16: ef_data.get("cfm") or "",
            18: ef_data.get("esp") or "",
            21: ef_data.get("rpm") or "",
            26: voltage,
        }
        
        for row, value in data_map.items():
            if value:
                self._safe_set_cell(ws, row, 14, value)
        
        print(f"  [OK] Populated EF sheet: {sheet_name}")
        return True
    
    def populate_heater(self, heater_data: Dict, sheet_name: str, block: int = 1) -> bool:
        """Populate an Electric Duct Heater sheet"""
        if sheet_name not in self.sheet_names:
            print(f"  No sheet found: {sheet_name}")
            return False
        
        ws = self.wb[sheet_name]
        
        # Use block1 or block2 mapping
        mapping = self.HEATER_MAPPING[f"block{block}"]
        
        for row, (label, col) in mapping.items():
            col_idx = ord(col) - ord('A') + 1
            
            if "Unit Number" in label:
                self._safe_set_cell(ws, row, col_idx, heater_data.get("tag", ""))
            elif "Location" in label:
                self._safe_set_cell(ws, row, col_idx, heater_data.get("location", ""))
            elif "CFM" in label:
                self._safe_set_cell(ws, row, col_idx, heater_data.get("cfm", ""))
            elif "Voltage" in label:
                self._safe_set_cell(ws, row, col_idx, heater_data.get("voltage", ""))
            elif "KW" in label:
                self._safe_set_cell(ws, row, col_idx, heater_data.get("kw", ""))
        
        print(f"  [OK] Populated Heater sheet: {sheet_name} (block {block})")
        return True
    
    def populate_all(self, extracted_data: Dict[str, List]) -> Dict[str, int]:
        """
        Populate all sheets with extracted data
        
        Args:
            extracted_data: Dict from LLM extractor with keys: fans, vavs, cracs, heaters
            
        Returns:
            Dict with count of populated sheets per type
        """
        stats = {"vavs": 0, "fans": 0, "heaters": 0, "cracs": 0}
        
        print("\nPopulating VAV sheets...")
        for vav in extracted_data.get("vavs", []):
            if self.populate_vav(vav):
                stats["vavs"] += 1
        
        print("\nPopulating Exhaust Fan sheets...")
        for ef in extracted_data.get("fans", []):
            if self.populate_ef(ef):
                stats["fans"] += 1
        
        print("\nPopulating Electric Duct Heater sheets...")
        heater_sheets = [s for s in self.sheet_names if "Electric Duct Heater" in s]
        heaters = extracted_data.get("heaters", [])
        
        # Pair heaters into sheets (2 per sheet)
        heater_idx = 0
        for sheet in heater_sheets:
            if heater_idx < len(heaters):
                self.populate_heater(heaters[heater_idx], sheet, block=1)
                stats["heaters"] += 1
                heater_idx += 1
            if heater_idx < len(heaters):
                self.populate_heater(heaters[heater_idx], sheet, block=2)
                stats["heaters"] += 1
                heater_idx += 1
        
        return stats
    
    def save(self, output_path: str = None):
        """Save the populated workbook"""
        output = output_path or self.template_path.replace(".xlsx", "_populated.xlsx")
        self.wb.save(output)
        print(f"\n[OK] Saved to: {output}")
        return output


def populate_excel_from_json(json_path: str, template_path: str, output_path: str = None) -> str:
    """
    Convenience function to populate Excel from extracted JSON data
    
    Args:
        json_path: Path to extracted_hvac_data.json
        template_path: Path to Boeing template Excel
        output_path: Optional output path
        
    Returns:
        Path to populated Excel file
    """
    with open(json_path) as f:
        data = json.load(f)
    
    populator = HVACExcelPopulator(template_path)
    stats = populator.populate_all(data)
    
    print(f"\nPopulation Summary:")
    print(f"  VAVs: {stats['vavs']}")
    print(f"  Fans: {stats['fans']}")
    print(f"  Heaters: {stats['heaters']}")
    
    return populator.save(output_path)


if __name__ == "__main__":
    # Test with sample data
    TEMPLATE = r"D:\SW\new project\Boeing Arlington R&D Setup.xlsx"
    JSON_DATA = r"D:\SW\new project\extracted_hvac_data.json"
    
    if Path(JSON_DATA).exists():
        output = populate_excel_from_json(JSON_DATA, TEMPLATE)
        print(f"Output: {output}")
    else:
        print(f"No extracted data found at {JSON_DATA}")
        print("Run llm_extractor.py first to extract data from PDF")
