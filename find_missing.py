"""
Find missing sheets and values between original and generated Excel
"""
import openpyxl

original = openpyxl.load_workbook(r'D:\SW\new project\Boeing Arlington R&D Setup.xlsx')
generated = openpyxl.load_workbook(r'D:\SW\new project\output\hvac_report.xlsx')

orig_sheets = set(original.sheetnames)
gen_sheets = set(generated.sheetnames)

missing = orig_sheets - gen_sheets

print("=" * 60)
print("MISSING SHEETS FROM GENERATED")
print("=" * 60)

# Categorize
categories = {
    'VAV': [],
    'EF': [],
    'Heaters': [],
    'Flow Meters': [],
    'Other': []
}

for s in sorted(missing):
    if s.startswith('VAVB'):
        categories['VAV'].append(s)
    elif s.startswith('EF'):
        categories['EF'].append(s)
    elif 'Heater' in s or 'Electric' in s:
        categories['Heaters'].append(s)
    elif 'Flow' in s:
        categories['Flow Meters'].append(s)
    else:
        categories['Other'].append(s)

for cat, sheets in categories.items():
    if sheets:
        print(f"\n{cat} ({len(sheets)}):")
        for s in sheets:
            print(f"  - {s}")

print(f"\nTotal missing: {len(missing)} sheets")
print(f"Generated has: {len(gen_sheets)} sheets")
print(f"Original has: {len(orig_sheets)} sheets")

# Check what fields are in original VAV but not in generated
print("\n" + "=" * 60)
print("MISSING FIELDS IN VAV SHEETS")
print("=" * 60)

orig_ws = original['VAVB5-01']
gen_ws = generated['VAVB5-01']

# Get all labels from original (column A/D)
orig_labels = set()
for i in range(1, 40):
    for j in [1, 4]:  # Columns A and D
        val = orig_ws.cell(i, j).value
        if val and isinstance(val, str):
            orig_labels.add(val.strip())

# Get all labels from generated
gen_labels = set()
for i in range(1, 25):
    val = gen_ws.cell(i, 1).value
    if val and isinstance(val, str):
        gen_labels.add(val.strip())

missing_fields = orig_labels - gen_labels
print("\nFields in Original but not in Generated:")
for f in sorted(missing_fields):
    if len(f) > 3:  # Skip short labels
        print(f"  - {f}")
