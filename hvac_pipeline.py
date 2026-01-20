"""
HVAC Data Extraction Pipeline - Complete Version
Uses improved extractor for accurate VAV data extraction
Generates clean Excel output (template-independent)
"""
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from typing import List

from extractors.improved_extractor import extract_vavs, VAVData


class HVACPipeline:
    """Complete HVAC extraction and Excel generation pipeline"""
    
    def __init__(
        self,
        pdf_path: str,
        output_path: str = "output/hvac_extracted.xlsx",
        job_number: int = 1168,
        project_name: str = "HVAC Project"
    ):
        self.pdf_path = pdf_path
        self.output_path = output_path
        self.job_number = job_number
        self.project_name = project_name
        self.vavs: List[VAVData] = []
        
    def extract(self):
        """Extract VAV data from PDF"""
        print("Extracting VAV data from PDF...")
        self.vavs = extract_vavs(self.pdf_path)
        print(f"  Found {len(self.vavs)} VAV units")
        
    def _create_vav_sheet(self, wb: Workbook, vav: VAVData):
        """Create a VAV sheet with extracted data"""
        ws = wb.create_sheet(title=vav.tag)
        
        # Styles
        bold = Font(bold=True)
        header_font = Font(bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "Fan Powered VAV Box Test Data"
        ws['A1'].font = header_font
        
        # Job info
        ws['A3'] = "Job Number:"
        ws['C3'] = self.job_number
        ws['E3'] = "Date:"
        ws['F3'] = datetime.now().strftime('%Y-%m-%d')
        
        ws['A4'] = "System:"
        ws['C4'] = vav.tag
        
        # Unit Information Section
        ws['A6'] = "UNIT INFORMATION"
        ws['A6'].font = bold
        
        ws['A7'] = "Unit Number"
        ws['C7'] = vav.tag
        
        ws['A8'] = "Location"
        ws['C8'] = vav.location or ""
        
        ws['A9'] = "Area Served"
        ws['C9'] = vav.area_served or ""
        
        ws['A10'] = "Inlet Size"
        ws['C10'] = vav.inlet_size
        
        # Air Measurements Section
        ws['A12'] = "AIR MEASUREMENTS"
        ws['A12'].font = bold
        ws['C12'] = "DESIGN"
        ws['C12'].font = bold
        ws['D12'] = "ACTUAL"
        ws['D12'].font = bold
        
        ws['A13'] = "Total Fan CFM"
        ws['C13'] = vav.total_cfm
        
        ws['A14'] = "Minimum CFM"
        ws['C14'] = vav.min_cfm
        
        ws['A15'] = "Maximum CFM"
        ws['C15'] = vav.max_cfm
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
    def _create_summary_sheet(self, wb: Workbook):
        """Create summary sheet with all VAV data"""
        ws = wb.active
        ws.title = "VAV Summary"
        
        # Headers
        headers = ["VAV Tag", "Total CFM", "Min CFM", "Max CFM", "Inlet Size", "Page"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, vav in enumerate(sorted(self.vavs, key=lambda v: v.tag), 2):
            ws.cell(row=row, column=1, value=vav.tag)
            ws.cell(row=row, column=2, value=vav.total_cfm)
            ws.cell(row=row, column=3, value=vav.min_cfm)
            ws.cell(row=row, column=4, value=vav.max_cfm)
            ws.cell(row=row, column=5, value=vav.inlet_size)
            ws.cell(row=row, column=6, value=vav.page + 1)
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
    def generate_excel(self):
        """Generate Excel workbook with extracted data"""
        print("\nGenerating Excel output...")
        
        wb = Workbook()
        
        # Create summary sheet
        self._create_summary_sheet(wb)
        
        # Create individual VAV sheets
        for vav in sorted(self.vavs, key=lambda v: v.tag):
            if vav.tag.startswith("VAVB"):  # Only VAVB5-XX sheets
                self._create_vav_sheet(wb, vav)
        
        # Save
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        print(f"  Saved to: {self.output_path}")
        
    def run(self):
        """Run complete pipeline"""
        print("=" * 60)
        print("HVAC EXTRACTION PIPELINE")
        print("=" * 60)
        print(f"PDF: {self.pdf_path}")
        print(f"Output: {self.output_path}")
        print()
        
        self.extract()
        self.generate_excel()
        
        print()
        print("=" * 60)
        print("COMPLETE!")
        print(f"  VAV Units Extracted: {len(self.vavs)}")
        print(f"  Output File: {self.output_path}")
        print("=" * 60)


def main():
    PDF_PATH = r"D:\SW\new project\Boeing R&D Drawings.pdf"
    OUTPUT_PATH = "output/hvac_extracted.xlsx"
    
    pipeline = HVACPipeline(
        pdf_path=PDF_PATH,
        output_path=OUTPUT_PATH,
        job_number=1168,
        project_name="Boeing Arlington R&D"
    )
    
    pipeline.run()


if __name__ == "__main__":
    main()
